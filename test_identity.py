# -*- coding: utf-8 -*-
"""验证：网页版 build_warehouse_workbook 的输出 == 原脚本逻辑的输出（同一份数据）。"""
import io, re, datetime, hashlib, sys
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

# ---- 原脚本的单仓库生成逻辑（从《赫眉盘点表12》逐行抠出，不改）----
def original_build(warehouse, data):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    column_widths = {'A':18,'B':67,'C':25,'D':16,'E':13,'F':10,'G':20,'H':25,'I':20,'J':20,'K':23,'L':20,'M':20}
    data = data.sort_values(by=["商品名称","品类","品牌"], ascending=[False,False,False])
    data['品牌'] = data['品牌'].apply(lambda x: re.sub(r'^\([0-9]+\)','',str(x)).strip())
    data["实际数量"]=""; data["有效期"]=""; data["备注"]=""
    wb = openpyxl.Workbook(); ws = wb.active
    example_row = ["举例","Lumi综合发酵果粉（固体饮料）","4712847540248","Lumi","79","10","","","","","","","",""]
    ws.append(example_row); ws['G1']='3+8'; ws['H1']='20251231'
    ws['G1'].alignment=Alignment(horizontal='center',vertical='center')
    ws['H1'].alignment=Alignment(horizontal='center',vertical='center')
    column_names=["商品编号","商品名称","商品条码","品牌","零售价","库存\r总数","实际数量","有效期","备注","清理计划","实际清理数量","剩余数量","品类"]
    ws.append(column_names)
    for _, row in data.iterrows():
        ws.append([row["商品编号"],row["商品名称"],row["商品条码"],row["品牌"],row["零售价"],row["库存总数"],"","","","","","",row["品类"]])
    for col,width in column_widths.items(): ws.column_dimensions[col].width=width
    default_font_size=20; name_column_font_size=20; default_row_height=40; header_row_height=50
    ws.row_dimensions[2].height=header_row_height
    for col in range(1,len(column_names)+1):
        cell=ws.cell(row=2,column=col)
        if col==6: cell.alignment=Alignment(vertical='center',horizontal='center',wrap_text=True)
        elif col==4: cell.alignment=Alignment(vertical='center',horizontal='center',wrap_text=True)
        elif col in [7,8,9]: cell.alignment=Alignment(vertical='center',horizontal='center',wrap_text=True)
        else: cell.alignment=Alignment(vertical='center',wrap_text=True)
        cell.border=thin_border
        cell.font=Font(size=20) if col in [10,11,12,13] else Font(size=default_font_size)
    for row in range(1,ws.max_row+1):
        if row==2: continue
        ws.row_dimensions[row].height=default_row_height
        for col in range(1,len(column_names)+1):
            cell=ws.cell(row=row,column=col); cell.border=thin_border
            cell.font=Font(size=name_column_font_size) if col==2 else Font(size=default_font_size)
            cell.alignment=Alignment(vertical='center')
            if col==1:
                cell.number_format='@'; cell.alignment=Alignment(vertical='center',horizontal='left')
                if isinstance(cell.value,(int,float)): cell.value=str(int(cell.value))
            elif col==2 and cell.value:
                text=str(cell.value); cell.alignment=Alignment(wrap_text=True,vertical='center',horizontal='left')
                chars_per_line=int(column_widths['B']*2); text_length=len(text)
                rows_needed=(text_length+chars_per_line-1)//chars_per_line
                row_height=max(default_row_height,min(rows_needed*default_row_height,80))
                ws.row_dimensions[row].height=row_height
            elif col==3:
                cell.number_format='@'; cell.alignment=Alignment(vertical='center',horizontal='left')
                if isinstance(cell.value,(int,float)): cell.value=str(int(cell.value))
            elif col==4: cell.alignment=Alignment(vertical='center',horizontal='center')
            elif col in [5,6]: cell.alignment=Alignment(vertical='center',horizontal='center')
            else: cell.alignment=Alignment(vertical='center',horizontal='left')
    yellow_fill=PatternFill(start_color="FFFF00",end_color="FFFF00",fill_type="solid")
    blue_fill=PatternFill(start_color="00B0F0",end_color="00B0F0",fill_type="solid")
    black_font=Font(color="000000")
    ws.cell(row=2,column=10).fill=yellow_fill; ws.cell(row=2,column=10).font=black_font
    ws.cell(row=2,column=11).fill=blue_fill; ws.cell(row=2,column=11).font=black_font
    ws.cell(row=2,column=12).fill=blue_fill; ws.cell(row=2,column=12).font=black_font
    ws.cell(row=2,column=13).font=black_font
    ws.freeze_panes='A3'
    ws.page_setup.paperSize=ws.PAPERSIZE_A4; ws.page_setup.orientation=ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage=True; ws.page_setup.fitToHeight=False; ws.page_setup.fitToWidth=1
    tomorrow=(datetime.datetime.now()+datetime.timedelta(days=1)).strftime('%Y%m%d')
    ws.oddHeader.left.text=f'&20盘点{warehouse} {tomorrow}'
    ws.oddHeader.center.text='&20第&[Page]页，共&[Pages]页'
    ws.oddHeader.right.text='&20有效期统计2027年12月31日前'
    ws.page_margins.left=0.1; ws.page_margins.right=0.1; ws.page_margins.top=0.4
    ws.page_margins.bottom=0.1; ws.page_margins.header=0.2; ws.page_margins.footer=0
    ws.print_title_rows='1:2'; ws.print_area=f'A1:I{ws.max_row}'
    for col in ['J','K','L','M']: ws.column_dimensions[col].hidden=True
    ws['G1'].alignment=Alignment(horizontal='center',vertical='center')
    ws['H1'].alignment=Alignment(horizontal='center',vertical='center')
    for col in ['J','K','L','M']: ws[f'{col}2'].font=Font(size=20)
    ws.sheet_view.zoomScale=50
    return wb

# ---- 构造样本数据 ----
sample = pd.DataFrame({
    "仓库名称":["A仓","A仓","B仓","A仓","B仓"],
    "商品编号":["0001","0002","0010","0003","0011"],
    "商品名称":["柔肤水 滋润型 200ml","精华液 抗皱","面膜 补水 25片装超长名称测试换行效果商品名称要够长才能触发多行高度逻辑","洁面乳","眼霜"],
    "商品条码":["6900000000001","6900000000002","6900000000010","6900000000003","6900000000011"],
    "品牌":["(12)Lumi","(3)Aqua","Lumi","(7)Aqua","Lumi"],
    "零售价":[79,129,49,59,199],
    "库存总数":[10,5,20,8,3],
    "品类":["护肤","护肤","面膜","清洁","护肤"],
})

from app import build_warehouse_workbook

def digest(wb):
    b=io.BytesIO(); wb.save(b); return b.getvalue()

all_ok=True
for wh, grp in sample.groupby("仓库名称"):
    wb_new,_ = build_warehouse_workbook(wh, grp.copy())
    wb_old   = original_build(wh, grp.copy())
    a=digest(wb_new); b=digest(wb_old)
    same = (a==b)
    all_ok &= same
    print(f"[{wh}] 网页版 {len(a)}B  vs  原逻辑 {len(b)}B  ->  {'✅ 字节完全一致' if same else '❌ 不一致'}")

print("\n总结：", "✅ 全部一致，网页版输出 == 原脚本输出" if all_ok else "❌ 存在差异")
sys.exit(0 if all_ok else 1)

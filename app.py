# -*- coding: utf-8 -*-
"""
赫眉 | HYMN 盘点表网页版
MAISON LUMIÈRE 风格 · Flask 后端
处理逻辑 100% 复用原脚本《赫眉盘点表12(2026.06.06).py》的 openpyxl 代码，
保证下载的 xlsx 与原脚本输出一致。
"""
import io
import re
import uuid
import zipfile
import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

from flask import Flask, request, jsonify, send_file, abort, Response

app = Flask(__name__)

# 处理结果暂存（本地单用户，内存即可）：token -> {"files": {name: bytes}, "name": 原文件名}
RESULTS = {}

REQUIRED_COLS = ["商品编号", "商品名称", "商品条码", "品牌", "零售价", "库存总数", "品类", "仓库名称"]


# ───────────────────────────────────────────────────────────────────
# 核心处理：与原脚本逐行一致，仅把「读文件/存文件夹」改为「读字节/存内存」
# ───────────────────────────────────────────────────────────────────
DEFAULT_HEADER_RIGHT = "有效期统计2027年12月31日前"


def build_warehouse_workbook(warehouse, data, header_right=DEFAULT_HEADER_RIGHT):
    """为单个仓库生成 openpyxl 工作簿（逻辑同原脚本），返回 (wb, preview_dict)。"""
    data = data.sort_values(
        by=["商品名称", "品类", "品牌"],
        ascending=[False, False, False]
    )

    data['品牌'] = data['品牌'].apply(lambda x: re.sub(r'^\([0-9]+\)', '', str(x)).strip())

    data["实际数量"] = ""
    data["有效期"] = ""
    data["备注"] = ""

    wb = openpyxl.Workbook()
    ws = wb.active

    # 1. 写入示例行
    example_row = ["举例", "Lumi综合发酵果粉（固体饮料）", "4712847540248", "Lumi", "79", "10", "", "", "", "", "", "", "", ""]
    ws.append(example_row)
    ws['G1'] = '3+8'
    ws['H1'] = '20251231'
    ws['G1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['H1'].alignment = Alignment(horizontal='center', vertical='center')

    # 2. 写入列名
    column_names = ["商品编号", "商品名称", "商品条码", "品牌", "零售价", "库存\r总数", "实际数量", "有效期", "备注",
                    "清理计划", "实际清理数量", "剩余数量", "品类"]
    ws.append(column_names)

    preview_rows = []

    # 3. 写入排序后的数据
    for _, row in data.iterrows():
        row_data = [
            row["商品编号"], row["商品名称"], row["商品条码"], row["品牌"],
            row["零售价"], row["库存总数"], "", "", "", "", "", "", row["品类"]
        ]
        ws.append(row_data)
        preview_rows.append([
            _s(row["商品编号"]), _s(row["商品名称"]), _s(row["商品条码"]),
            _s(row["品牌"]), _s(row["零售价"]), _s(row["库存总数"]), _s(row["品类"]),
        ])

    column_widths = {
        'A': 18, 'B': 67, 'C': 25, 'D': 16, 'E': 13, 'F': 10, 'G': 20,
        'H': 25, 'I': 20, 'J': 20, 'K': 23, 'L': 20, 'M': 20,
    }

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    default_font_size = 20
    name_column_font_size = 20
    default_row_height = 40
    header_row_height = 50

    ws.row_dimensions[2].height = header_row_height
    for col in range(1, len(column_names) + 1):
        cell = ws.cell(row=2, column=col)
        if col == 6:
            cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
        elif col == 4:
            cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
        elif col in [7, 8, 9]:
            cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
        else:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border
        if col in [10, 11, 12, 13]:
            cell.font = Font(size=20)
        else:
            cell.font = Font(size=default_font_size)

    for row in range(1, ws.max_row + 1):
        if row == 2:
            continue
        ws.row_dimensions[row].height = default_row_height
        for col in range(1, len(column_names) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col == 2:
                cell.font = Font(size=name_column_font_size)
            else:
                cell.font = Font(size=default_font_size)
            cell.alignment = Alignment(vertical='center')
            if col == 1:
                cell.number_format = '@'
                cell.alignment = Alignment(vertical='center', horizontal='left')
                if isinstance(cell.value, (int, float)):
                    cell.value = str(int(cell.value))
            elif col == 2 and cell.value:
                text = str(cell.value)
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                chars_per_line = int(column_widths['B'] * 2)
                text_length = len(text)
                rows_needed = (text_length + chars_per_line - 1) // chars_per_line
                row_height = max(default_row_height, min(rows_needed * default_row_height, 80))
                ws.row_dimensions[row].height = row_height
            elif col == 3:
                cell.number_format = '@'
                cell.alignment = Alignment(vertical='center', horizontal='left')
                if isinstance(cell.value, (int, float)):
                    cell.value = str(int(cell.value))
            elif col == 4:
                cell.alignment = Alignment(vertical='center', horizontal='center')
            elif col in [5, 6]:
                cell.alignment = Alignment(vertical='center', horizontal='center')
            else:
                cell.alignment = Alignment(vertical='center', horizontal='left')

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    black_font = Font(color="000000")

    ws.cell(row=2, column=10).fill = yellow_fill
    ws.cell(row=2, column=10).font = black_font
    ws.cell(row=2, column=11).fill = blue_fill
    ws.cell(row=2, column=11).font = black_font
    ws.cell(row=2, column=12).fill = blue_fill
    ws.cell(row=2, column=12).font = black_font
    ws.cell(row=2, column=13).font = black_font

    ws.freeze_panes = 'A3'

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = False
    ws.page_setup.fitToWidth = 1

    tomorrow = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%Y%m%d')
    ws.oddHeader.left.text = f'&20盘点{warehouse} {tomorrow}'
    ws.oddHeader.center.text = '&20第&[Page]页，共&[Pages]页'
    ws.oddHeader.right.text = '&20' + header_right

    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.1
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0

    ws.print_title_rows = '1:2'
    ws.print_area = f'A1:I{ws.max_row}'

    for col in ['J', 'K', 'L', 'M']:
        ws.column_dimensions[col].hidden = True

    ws['G1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['H1'].alignment = Alignment(horizontal='center', vertical='center')
    for col in ['J', 'K', 'L', 'M']:
        ws[f'{col}2'].font = Font(size=20)
    ws.sheet_view.zoomScale = 50

    preview = {
        "warehouse": str(warehouse),
        "count": len(preview_rows),
        "rows": preview_rows,
    }
    return wb, preview


def _s(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def process_excel(file_bytes, header_right=DEFAULT_HEADER_RIGHT):
    bio = io.BytesIO(file_bytes)
    df = pd.read_excel(bio, dtype={'商品编号': str}, engine='openpyxl')
    df['商品编号'] = df['商品编号'].astype(str)

    if "仓库名称" not in df.columns:
        raise ValueError("未找到 '仓库名称' 列，请检查文件格式")

    missing = [c for c in ["商品名称", "商品条码", "品牌", "零售价", "库存总数", "品类"] if c not in df.columns]
    if missing:
        raise ValueError("缺少必要的列：" + "、".join(missing))

    grouped_data = df.groupby("仓库名称")

    files = {}
    previews = []
    for warehouse, data in grouped_data:
        wb, preview = build_warehouse_workbook(warehouse, data, header_right)
        out = io.BytesIO()
        wb.save(out)
        files[f"{warehouse}.xlsx"] = out.getvalue()
        previews.append(preview)

    return files, previews


# ───────────────────────────────────────────────────────────────────
# 路由
# ───────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return Response(INDEX_HTML, mimetype="text/html")


@app.route("/process", methods=["POST"])
def process():
    if "file" not in request.files:
        return jsonify({"error": "未收到文件"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "未选择文件"}), 400
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "请上传 .xlsx 或 .xls 文件"}), 400

    header_right = (request.form.get("header_right") or "").strip() or DEFAULT_HEADER_RIGHT

    try:
        files, previews = process_excel(f.read(), header_right)
    except Exception as e:  # noqa: BLE001
        return jsonify({"error": f"处理失败：{e}"}), 400

    if not files:
        return jsonify({"error": "没有可拆分的数据（仓库分组为空）"}), 400

    token = uuid.uuid4().hex
    RESULTS[token] = {"files": files, "name": f.filename}

    total = sum(p["count"] for p in previews)
    header_date = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%Y%m%d')
    return jsonify({
        "token": token,
        "source": f.filename,
        "warehouse_count": len(files),
        "total_rows": total,
        "header_date": header_date,
        "header_right": header_right,
        "zip_name": f"盘点表格汇总_{header_date}.zip",
        "previews": previews,
    })


@app.route("/download/<token>")
def download_zip(token):
    item = RESULTS.get(token)
    if not item:
        abort(404)
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in item["files"].items():
            zf.writestr(name, data)
    mem.seek(0)
    stamp = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%Y%m%d')
    return send_file(mem, mimetype="application/zip", as_attachment=True,
                     download_name=f"盘点表格汇总_{stamp}.zip")


@app.route("/download/<token>/<path:name>")
def download_one(token, name):
    item = RESULTS.get(token)
    if not item or name not in item["files"]:
        abort(404)
    return send_file(io.BytesIO(item["files"][name]),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=name)


# ───────────────────────────────────────────────────────────────────
# 前端（MAISON LUMIÈRE · 3D 质感 · 左右分栏 · A4 打印预览）
# ───────────────────────────────────────────────────────────────────
INDEX_HTML = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>赫眉 · 盘点表生成</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@500;600;700;800&family=DM+Sans:wght@400;500;600;700&family=Noto+Serif+SC:wght@500;600;700&family=Noto+Sans+SC:wght@700;900&display=swap" rel="stylesheet">
<style>
  :root{
    --ivory:#FBF8F2; --ivory-2:#F1E8DA; --orange:#E8590C; --orange-soft:#F5772E; --orange-deep:#C9490A;
    --gold:#B9974C; --gold-lt:#E7CE97; --ink:#221D17; --muted:#8C8073; --line:rgba(34,29,23,.10);
  }
  *{box-sizing:border-box;margin:0;padding:0}
  html{scroll-behavior:smooth}
  body{
    font-family:'DM Sans','Noto Serif SC',-apple-system,sans-serif;
    color:var(--ink); min-height:100vh; overflow-x:hidden; position:relative;
    display:flex; flex-direction:column; align-items:center;
    background:
      radial-gradient(1100px 560px at 80% -10%, rgba(232,89,12,.08), transparent 60%),
      radial-gradient(900px 480px at 5% 108%, rgba(185,151,76,.12), transparent 55%),
      linear-gradient(160deg,#FFFDF8 0%, var(--ivory) 42%, var(--ivory-2) 100%);
    -webkit-font-smoothing:antialiased;
  }
  .orb{position:fixed;border-radius:50%;filter:blur(42px);opacity:.5;z-index:0;pointer-events:none;
    animation:drift 18s ease-in-out infinite}
  .orb.a{width:320px;height:320px;top:-90px;right:6%;
    background:radial-gradient(circle at 30% 30%, rgba(245,119,46,.55), rgba(232,89,12,0) 70%)}
  .orb.b{width:400px;height:400px;bottom:-120px;left:-60px;animation-delay:-7s;
    background:radial-gradient(circle at 40% 40%, rgba(231,206,151,.6), rgba(185,151,76,0) 70%)}
  .orb.c{width:220px;height:220px;top:40%;left:50%;animation-delay:-12s;opacity:.3;
    background:radial-gradient(circle at 40% 40%, rgba(245,119,46,.4), transparent 70%)}
  @keyframes drift{0%,100%{transform:translate(0,0)}50%{transform:translate(26px,-32px)}}
  /* 高科技粒子星座背景 + 极光流光 */
  #bg{position:fixed;inset:0;z-index:0;pointer-events:none}
  .aurora{position:fixed;inset:-20%;z-index:0;pointer-events:none;opacity:.55;mix-blend-mode:multiply;
    background:
      radial-gradient(40% 30% at 20% 30%, rgba(232,89,12,.16), transparent 60%),
      radial-gradient(36% 28% at 80% 40%, rgba(185,151,76,.20), transparent 60%),
      radial-gradient(44% 34% at 55% 80%, rgba(245,119,46,.12), transparent 60%);
    filter:blur(20px);animation:aur 22s ease-in-out infinite}
  @keyframes aur{0%,100%{transform:translate(0,0) rotate(0deg) scale(1)}
    33%{transform:translate(-3%,2%) rotate(4deg) scale(1.05)}
    66%{transform:translate(3%,-2%) rotate(-3deg) scale(1.02)}}
  .grid-veil{position:fixed;inset:0;z-index:0;pointer-events:none;opacity:.4;
    background-image:linear-gradient(rgba(34,29,23,.04) 1px,transparent 1px),
      linear-gradient(90deg,rgba(34,29,23,.04) 1px,transparent 1px);
    background-size:46px 46px;
    -webkit-mask:radial-gradient(circle at 50% 30%,#000,transparent 78%);
    mask:radial-gradient(circle at 50% 30%,#000,transparent 78%)}

  .wrap{position:relative;z-index:1;width:100%;max-width:1380px;padding:40px 26px;perspective:1700px;
    min-height:100vh;display:flex;flex-direction:column;justify-content:center}

  /* 页眉编辑条 */
  .ppbar{display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-bottom:14px;
    background:linear-gradient(180deg,rgba(255,255,255,.9),rgba(251,244,234,.85));
    border:1px solid var(--line);border-radius:14px;padding:11px 14px;
    box-shadow:0 1px 0 rgba(255,255,255,.9) inset,0 14px 28px -22px rgba(34,29,23,.4)}
  .ppbar .pl{font-size:13px;font-weight:600;color:var(--ink);white-space:nowrap;display:flex;align-items:center;gap:6px}
  .ppbar input{flex:1;min-width:180px;font-family:inherit;font-size:13.5px;color:var(--ink);
    border:1px solid var(--line);border-radius:10px;padding:9px 12px;background:#fff;transition:.16s}
  .ppbar input:focus{outline:none;border-color:var(--orange);box-shadow:0 0 0 3px rgba(232,89,12,.12)}
  .ppbar .mini{appearance:none;border:none;cursor:pointer;font-family:inherit;font-weight:700;font-size:13px;
    letter-spacing:.08em;color:#fff;padding:9px 18px;border-radius:10px;
    background:linear-gradient(180deg,#FA8246,var(--orange));white-space:nowrap;
    box-shadow:0 10px 20px -10px rgba(232,89,12,.6)}
  .ppbar .mini:hover{filter:brightness(1.05);transform:translateY(-1px)}
  .ppbar .mini:active{transform:translateY(1px)}

  /* ── 顶部品牌（紧凑横排，省竖向空间） ── */
  .brand{display:flex;align-items:center;justify-content:center;gap:20px;margin-bottom:26px;flex-wrap:wrap}
  .saturn{width:74px;height:74px;position:relative;transform-style:preserve-3d;flex:none;
    animation:floaty 6s ease-in-out infinite}
  @keyframes floaty{0%,100%{transform:translateY(0)}50%{transform:translateY(-8px)}}
  .saturn .glow{position:absolute;inset:-12px;border-radius:50%;
    background:radial-gradient(circle, rgba(232,89,12,.32), transparent 65%);filter:blur(7px)}
  .saturn .planet{position:absolute;inset:21px;border-radius:50%;
    background:radial-gradient(circle at 34% 28%, #FFCBA0 0%, #F5772E 38%, var(--orange) 64%, var(--orange-deep) 100%);
    box-shadow:inset -6px -7px 14px rgba(120,40,0,.5),inset 5px 5px 11px rgba(255,210,170,.7),0 10px 22px -6px rgba(232,89,12,.6)}
  .saturn .planet::after{content:"";position:absolute;top:18%;left:22%;width:30%;height:22%;
    border-radius:50%;background:radial-gradient(circle,rgba(255,255,255,.85),transparent 70%);filter:blur(2px)}
  .saturn .ring{position:absolute;inset:4px;border-radius:50%;transform-style:preserve-3d;border:2.5px solid transparent;
    border-top-color:var(--gold-lt);border-bottom-color:var(--gold);transform:rotateX(74deg) rotateZ(0deg);
    box-shadow:0 0 12px rgba(185,151,76,.4);animation:spin 7s linear infinite}
  .saturn .ring.r2{inset:11px;border-top-color:rgba(231,206,151,.6);border-bottom-color:rgba(185,151,76,.7);
    animation-duration:5s;animation-direction:reverse}
  @keyframes spin{to{transform:rotateX(74deg) rotateZ(360deg)}}
  .brand .txt{text-align:left}
  .eyebrow{font-family:'Playfair Display',serif;letter-spacing:.46em;font-size:11px;color:var(--gold);
    text-transform:uppercase;font-weight:600;margin-bottom:7px;text-shadow:0 1px 0 rgba(255,255,255,.7);
    display:inline-flex;align-items:center;gap:11px}
  .eyebrow::before,.eyebrow::after{content:"";width:22px;height:1px;
    background:linear-gradient(90deg,transparent,rgba(185,151,76,.85),transparent)}
  h1{font-family:'Playfair Display','Noto Serif SC',serif;font-weight:800;font-size:40px;letter-spacing:.08em;
    background:linear-gradient(180deg,#3A322A,#1A1510);-webkit-background-clip:text;background-clip:text;
    -webkit-text-fill-color:transparent;line-height:1.12}
  .sub{letter-spacing:.36em;font-size:10.5px;color:var(--muted);text-transform:uppercase;margin-top:8px}

  /* ── 左右分栏 ── */
  .grid{display:grid;grid-template-columns:minmax(330px,380px) 1fr;gap:26px;align-items:start}
  @media(max-width:980px){.grid{grid-template-columns:1fr}}

  .card{position:relative;background:linear-gradient(180deg,rgba(255,255,255,.86),rgba(255,255,255,.7));
    backdrop-filter:blur(18px) saturate(1.3);-webkit-backdrop-filter:blur(18px) saturate(1.3);
    border:1px solid rgba(255,255,255,.85);border-radius:24px;padding:30px;
    box-shadow:0 1px 0 rgba(255,255,255,.9) inset,0 40px 80px -34px rgba(34,29,23,.34),0 14px 30px -22px rgba(34,29,23,.3);
    transition:transform .4s cubic-bezier(.2,.8,.2,1), box-shadow .4s}
  .card::before{content:"";position:absolute;inset:0;border-radius:24px;padding:1px;pointer-events:none;
    background:linear-gradient(160deg,rgba(255,255,255,.9),transparent 40%,rgba(185,151,76,.25));
    -webkit-mask:linear-gradient(#000 0 0) content-box,linear-gradient(#000 0 0);
    -webkit-mask-composite:xor;mask-composite:exclude}
  .col-left{position:sticky;top:24px}
  .card h2{font-family:'Noto Serif SC',serif;font-size:21px;font-weight:600;margin-bottom:5px}
  .card .hint{color:var(--muted);font-size:13.5px;line-height:1.7;margin-bottom:22px}

  .drop{position:relative;border:1.6px dashed rgba(232,89,12,.4);border-radius:18px;
    background:linear-gradient(180deg,#FFFDFA,#FBF4EA);padding:40px 20px;text-align:center;cursor:pointer;
    transition:.22s;box-shadow:inset 0 2px 10px rgba(34,29,23,.05), inset 0 -2px 0 rgba(255,255,255,.8)}
  .drop:hover,.drop.over{border-color:var(--orange);background:#FFF6EE;
    box-shadow:inset 0 0 0 3px rgba(232,89,12,.07), 0 18px 40px -26px rgba(232,89,12,.5)}
  .drop .ico{font-size:28px;margin-bottom:12px;display:inline-block;animation:floaty 4s ease-in-out infinite}
  .drop .big{font-size:16px;font-weight:600;margin-bottom:6px}
  .drop .small{color:var(--muted);font-size:12.5px}
  .file-tag{display:inline-flex;align-items:center;gap:9px;margin-top:16px;background:#fff;border:1px solid var(--line);
    border-radius:999px;padding:8px 16px;font-size:13.5px;font-weight:500;box-shadow:0 8px 20px -12px rgba(34,29,23,.3)}
  .file-tag .dot{width:8px;height:8px;border-radius:50%;background:var(--orange);box-shadow:0 0 0 4px rgba(232,89,12,.15)}

  .row{display:flex;gap:12px;margin-top:22px;flex-wrap:wrap}
  .btn{appearance:none;border:none;cursor:pointer;font-family:inherit;font-weight:700;font-size:15px;
    letter-spacing:.16em;padding:15px 26px;border-radius:14px;transition:.18s;position:relative}
  .btn-primary{flex:1;min-width:180px;color:#fff;
    background:linear-gradient(180deg,#FA8246 0%, var(--orange) 52%, var(--orange-deep) 100%);
    box-shadow:0 1px 0 rgba(255,210,180,.7) inset,0 -3px 8px rgba(150,50,0,.3) inset,
      0 18px 34px -12px rgba(232,89,12,.7),0 6px 12px -6px rgba(232,89,12,.5)}
  .btn-primary:hover{transform:translateY(-2px);filter:brightness(1.05)}
  .btn-primary:active{transform:translateY(1px)}
  .btn-primary:disabled{opacity:.5;cursor:not-allowed;transform:none;filter:saturate(.6)}
  .btn-ghost{background:rgba(255,255,255,.7);border:1px solid var(--line);color:var(--ink);
    box-shadow:0 10px 24px -16px rgba(34,29,23,.4)}
  .btn-ghost:hover{border-color:var(--orange);color:var(--orange);transform:translateY(-1px)}

  /* 预览空状态 */
  .empty{display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;
    min-height:420px;color:var(--muted);gap:14px}
  .empty .ei{width:64px;height:64px;border-radius:18px;display:flex;align-items:center;justify-content:center;
    font-size:30px;background:linear-gradient(180deg,#FFF,#F4EADB);border:1px solid var(--line);
    box-shadow:0 16px 30px -20px rgba(34,29,23,.4)}
  .empty .et{font-size:15px;font-weight:600;color:var(--ink)}
  .empty .es{font-size:13px}

  #rbody{display:none}
  .stat-row{display:flex;gap:14px;flex-wrap:wrap;margin-bottom:20px}
  .stat{flex:1;min-width:130px;border-radius:16px;padding:16px 18px;
    background:linear-gradient(180deg,rgba(255,255,255,.9),rgba(251,244,234,.9));border:1px solid rgba(255,255,255,.8);
    box-shadow:0 1px 0 rgba(255,255,255,.9) inset,0 20px 40px -28px rgba(34,29,23,.35)}
  .stat .num{font-family:'Playfair Display',serif;font-size:30px;font-weight:800;color:var(--orange);line-height:1}
  .stat .lab{color:var(--muted);font-size:12.5px;letter-spacing:.06em;margin-top:5px}
  .tabs{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:16px;max-height:96px;overflow:auto}
  .tab{padding:8px 16px;border-radius:999px;border:1px solid var(--line);background:rgba(255,255,255,.7);
    cursor:pointer;font-size:13.5px;font-weight:600;transition:.16s;white-space:nowrap}
  .tab:hover{border-color:var(--orange);transform:translateY(-1px)}
  .tab.active{background:linear-gradient(180deg,#2E2820,var(--ink));color:#fff;border-color:var(--ink);
    box-shadow:0 12px 24px -14px rgba(34,29,23,.6)}
  .tab .n{opacity:.6;font-size:11.5px;margin-left:6px}

  .stage{background:radial-gradient(900px 300px at 50% 0%, rgba(34,29,23,.06), transparent 70%),
      linear-gradient(180deg,#EFE7DA,#E6DBC9);border-radius:18px;padding:28px 0;height:560px;overflow:auto;
    box-shadow:inset 0 2px 20px rgba(34,29,23,.12)}
  .pp-meta{text-align:center;color:var(--muted);font-size:11.5px;letter-spacing:.14em;text-transform:uppercase;margin-bottom:18px}
  .sheet{width:720px;max-width:94%;margin:0 auto 26px;background:#fff;color:#000;padding:26px 26px 34px;border-radius:3px;
    position:relative;box-shadow:0 2px 4px rgba(0,0,0,.16),0 30px 60px -24px rgba(0,0,0,.4);transition:transform .3s}
  .sheet:hover{transform:translateY(-3px)}
  .sheet-hd{display:flex;justify-content:space-between;align-items:flex-end;gap:10px;font-size:11px;color:#333;
    margin-bottom:9px;font-family:'DM Sans',sans-serif}
  .sheet-hd .c{text-align:center}
  .ptable{border-collapse:collapse;width:100%;table-layout:fixed;font-size:10.5px;color:#000;
    font-family:'DM Sans','Noto Serif SC',sans-serif}
  .ptable td,.ptable th{border:1px solid #000;padding:4px 5px;line-height:1.25;overflow:hidden;
    text-overflow:ellipsis;vertical-align:middle}
  .ptable thead th{font-weight:600;background:#fff;white-space:pre-line}
  .ptable .name{text-align:left;white-space:normal;word-break:break-all}
  .ta-c{text-align:center}.ta-l{text-align:left}

  .err{display:none;background:#FDECEC;border:1px solid #F3C6C6;color:#B23A3A;border-radius:12px;
    padding:13px 16px;font-size:13.5px;margin-top:14px}
  .spin{display:inline-block;width:15px;height:15px;border:2px solid rgba(255,255,255,.4);border-top-color:#fff;
    border-radius:50%;animation:spin2 .7s linear infinite;vertical-align:-2px;margin-right:8px}
  @keyframes spin2{to{transform:rotate(360deg)}}
  .dl-list{margin-top:16px;display:flex;flex-direction:column;gap:8px;max-height:170px;overflow:auto}
  .dl-item{display:flex;justify-content:space-between;align-items:center;background:rgba(255,255,255,.75);
    border:1px solid var(--line);border-radius:12px;padding:11px 16px;font-size:13.5px}
  .dl-item a{color:var(--orange);text-decoration:none;font-weight:700;cursor:pointer}
  .dl-item a:hover{text-decoration:underline}

  /* 下载提示 toast */
  .toast{position:fixed;left:50%;bottom:30px;transform:translateX(-50%) translateY(20px);z-index:50;
    background:linear-gradient(180deg,#2E2820,var(--ink));color:#fff;border-radius:14px;padding:15px 22px;
    box-shadow:0 24px 50px -18px rgba(34,29,23,.6);opacity:0;pointer-events:none;transition:.32s;max-width:88vw}
  .toast.show{opacity:1;transform:translateX(-50%) translateY(0)}
  .toast .tt{font-weight:700;font-size:14px;margin-bottom:4px;display:flex;align-items:center;gap:8px}
  .toast .tt .ck{color:#7CE3A6}
  .toast .td{font-size:12.5px;color:#D9CFC2;line-height:1.6}
  .toast .td b{color:#FFD9A8;font-weight:600}

  /* 顶部 masthead：左 logo / 中标题 */
  .masthead{display:grid;grid-template-columns:170px 1fr 170px;align-items:center;gap:16px;
    margin-bottom:30px;padding-bottom:22px;border-bottom:1px solid var(--line);position:relative}
  .masthead::after{content:"";position:absolute;left:0;right:0;bottom:-1px;height:1px;
    background:linear-gradient(90deg,transparent,rgba(185,151,76,.5),transparent)}
  .mark{width:158px;opacity:.95;filter:drop-shadow(0 6px 14px rgba(34,29,23,.1));transition:.25s}
  .mark:hover{opacity:1;transform:translateY(-1px)}
  .mark svg{display:block;width:100%;height:auto}
  .masthead .hero{justify-self:center;margin-bottom:0;flex-direction:column;gap:10px;text-align:center}
  @media(max-width:880px){.masthead{grid-template-columns:1fr;justify-items:center;gap:18px}
    .masthead .spacer{display:none}}
  .footlogo{width:200px;margin:0 auto 18px;opacity:.62}
  .footlogo svg{display:block;width:100%;height:auto}

  /* 空状态：虚化 A4 样张 */
  .ghost{width:100%;max-width:380px;margin:6px auto 0;background:#fff;border-radius:4px;padding:18px 18px 22px;
    box-shadow:0 2px 4px rgba(0,0,0,.06),0 24px 48px -28px rgba(34,29,23,.35);opacity:.62;
    -webkit-mask:linear-gradient(#000 60%,transparent);mask:linear-gradient(#000 60%,transparent)}
  .ghost .gh{display:flex;justify-content:space-between;font-size:9px;color:#bbb;margin-bottom:8px}
  .ghost table{width:100%;border-collapse:collapse}
  .ghost td{border:1px solid #e2e2e2;height:15px}
  .ghost .gbar{height:7px;border-radius:3px;background:#eee;margin:3px 4px}

  .foot{text-align:center;color:var(--muted);letter-spacing:.32em;font-size:10.5px;text-transform:uppercase;margin-top:30px}
  .foot::before{content:"";display:block;width:46px;height:1px;
    background:linear-gradient(90deg,transparent,var(--gold),transparent);margin:0 auto 14px}
</style>
</head>
<body>
<canvas id="bg"></canvas>
<div class="wrap">
  <header class="masthead">
    <div class="mark" id="topbrand"></div>
    <div class="brand hero">
      <div class="saturn"><div class="glow"></div><div class="ring"></div><div class="ring r2"></div><div class="planet"></div></div>
      <div class="txt">
        <div class="eyebrow">Maison Lumière</div>
        <h1>盘点表生成</h1>
        <div class="sub">Inventory Sheet Atelier</div>
      </div>
    </div>
    <div class="spacer"></div>
  </header>

  <div class="grid">
    <!-- 左：上传 -->
    <div class="col-left">
      <div class="card" id="upload-card">
        <h2>上传商品清单</h2>
        <div class="drop" id="drop">
          <div class="ico">📄</div>
          <div class="big">点击选择表格，或拖拽到此处</div>
          <div class="small">支持 .xlsx / .xls</div>
          <div id="filetag"></div>
        </div>
        <input type="file" id="file" accept=".xlsx,.xls" hidden>
        <div class="row"><button class="btn btn-primary" id="go" disabled>开 始 处 理</button></div>
        <div class="err" id="err"></div>
      </div>
    </div>

    <!-- 右：预览 -->
    <div class="card" id="result">
      <h2>处理结果 · 打印预览</h2>
      <div class="empty" id="empty">
        <div class="et">上传后在此预览打印效果</div>
        <div class="ghost">
          <div class="gh"><span>盘点 · 仓库 · 日期</span><span>有效期统计…前</span></div>
          <table><tbody>
            <tr><td colspan="5"><div class="gbar" style="width:60%"></div></td></tr>
            <tr><td></td><td></td><td></td><td></td><td></td></tr>
            <tr><td></td><td></td><td></td><td></td><td></td></tr>
            <tr><td></td><td></td><td></td><td></td><td></td></tr>
            <tr><td></td><td></td><td></td><td></td><td></td></tr>
            <tr><td></td><td></td><td></td><td></td><td></td></tr>
            <tr><td></td><td></td><td></td><td></td><td></td></tr>
          </tbody></table>
        </div>
      </div>
      <div id="rbody">
        <div class="hint" id="src-hint" style="margin-bottom:16px"></div>
        <div class="stat-row">
          <div class="stat"><div class="num" id="st-wh">0</div><div class="lab">店铺</div></div>
          <div class="stat"><div class="num" id="st-rows">0</div><div class="lab">商品行数</div></div>
          <div class="stat"><div class="num" id="st-pages">0</div><div class="lab">表格页数</div></div>
        </div>
        <div class="tabs" id="tabs"></div>
        <div class="ppbar">
          <span class="pl">📅 右上角有效期</span>
          <input id="hdr" value="有效期统计2027年12月31日前" spellcheck="false">
          <button class="mini" id="applyhdr">应用并更新</button>
        </div>
        <div class="stage" id="stage"></div>
        <div class="row">
          <button class="btn btn-primary" id="dlzip">下 载 全 部（ZIP）</button>
          <button class="btn btn-ghost" id="again">重新选择</button>
        </div>
        <div class="dl-list" id="dllist"></div>
      </div>
    </div>
  </div>

  <div class="foot">Crafted for Excellence</div>
</div>

<div class="toast" id="toast"></div>

<script>
const COLS = [
  {name:"商品编号", w:18, cls:"ta-l"},
  {name:"商品名称", w:67, cls:"name"},
  {name:"商品条码", w:25, cls:"ta-l"},
  {name:"品牌",     w:16, cls:"ta-c"},
  {name:"零售价",   w:13, cls:"ta-c"},
  {name:"库存\n总数", w:10, cls:"ta-c"},
  {name:"实际数量", w:20, cls:"ta-c"},
  {name:"有效期",   w:25, cls:"ta-c"},
  {name:"备注",     w:20, cls:"ta-l"},
];
const WSUM = COLS.reduce((s,c)=>s+c.w,0);
const EXAMPLE = ["举例","Lumi综合发酵果粉（固体饮料）","4712847540248","Lumi","79","10","3+8","20251231",""];
const ROWS_PER_PAGE = 20;

let chosen=null, data=null, active=0, toastTimer=null;
const $=s=>document.querySelector(s);
const drop=$('#drop'),fileInput=$('#file'),go=$('#go'),err=$('#err');

drop.onclick=()=>fileInput.click();
['dragover','dragenter'].forEach(e=>drop.addEventListener(e,ev=>{ev.preventDefault();drop.classList.add('over')}));
['dragleave','drop'].forEach(e=>drop.addEventListener(e,ev=>{ev.preventDefault();drop.classList.remove('over')}));
drop.addEventListener('drop',ev=>{if(ev.dataTransfer.files.length)setFile(ev.dataTransfer.files[0])});
fileInput.onchange=()=>{if(fileInput.files.length)setFile(fileInput.files[0])};

function setFile(f){chosen=f;go.disabled=false;err.style.display='none';
  $('#filetag').innerHTML='<span class="file-tag"><span class="dot"></span>'+esc(f.name)+'</span>';}

async function runProcess(){
  if(!chosen){ err.textContent='请先选择表格文件'; err.style.display='block'; return; }
  go.disabled=true; const old=go.innerHTML; go.innerHTML='<span class="spin"></span>处理中…'; err.style.display='none';
  try{
    const fd=new FormData(); fd.append('file',chosen);
    fd.append('header_right', ($('#hdr').value||'').trim());
    const r=await fetch('/process',{method:'POST',body:fd});
    const j=await r.json();
    if(!r.ok) throw new Error(j.error||'处理失败');
    data=j; render();
  }catch(e){ err.textContent=e.message; err.style.display='block'; }
  finally{ go.disabled=false; go.innerHTML=old; }
}
go.onclick=runProcess;

function render(){
  $('#empty').style.display='none';
  $('#rbody').style.display='block';
  $('#src-hint').textContent='来源：'+data.source+'　·　拆分日期 '+data.header_date;
  $('#st-wh').textContent=data.warehouse_count;
  $('#st-rows').textContent=data.total_rows;
  const tabs=$('#tabs'); tabs.innerHTML=''; active=0;
  data.previews.forEach((p,i)=>{
    const t=document.createElement('div'); t.className='tab'+(i===0?' active':'');
    t.innerHTML=esc(p.warehouse)+'<span class="n">'+p.count+'</span>';
    t.onclick=()=>{active=i;[...tabs.children].forEach(c=>c.classList.remove('active'));t.classList.add('active');drawSheets()};
    tabs.appendChild(t);
  });
  const dl=$('#dllist'); dl.innerHTML='';
  data.previews.forEach(p=>{
    const name=p.warehouse+'.xlsx';
    const it=document.createElement('div'); it.className='dl-item';
    it.innerHTML='<span>'+esc(name)+'　·　'+p.count+' 行</span><a>下载</a>';
    it.querySelector('a').onclick=()=>downloadFile('/download/'+data.token+'/'+encodeURIComponent(name), name);
    dl.appendChild(it);
  });
  drawSheets();
}

function colgroup(){return '<colgroup>'+COLS.map(c=>'<col style="width:'+(c.w/WSUM*100).toFixed(2)+'%">').join('')+'</colgroup>';}
function dataRow(r){
  const v=[r[0],r[1],r[2],r[3],r[4],r[5],'','',''];
  return '<tr>'+COLS.map((c,i)=>'<td class="'+c.cls+'">'+esc(v[i]||'')+'</td>').join('')+'</tr>';
}
function titleRows(){
  return '<thead>'
    + '<tr>'+COLS.map((c,i)=>'<th class="'+c.cls.replace('name','ta-l')+'" style="font-weight:400">'+esc(EXAMPLE[i]||'')+'</th>').join('')+'</tr>'
    + '<tr>'+COLS.map(c=>'<th class="'+c.cls.replace('name','ta-l')+'">'+esc(c.name)+'</th>').join('')+'</tr>'
    + '</thead>';
}
function drawSheets(){
  const p=data.previews[active];
  const rows=p.rows;
  const pages=Math.max(1,Math.ceil(rows.length/ROWS_PER_PAGE));
  $('#st-pages').textContent=pages;
  let html='<div class="pp-meta">A4 · 纵向 · 每页重复表头（共 '+pages+' 页）</div>';
  for(let pi=0; pi<pages; pi++){
    const slice=rows.slice(pi*ROWS_PER_PAGE,(pi+1)*ROWS_PER_PAGE);
    html+='<div class="sheet">'
      +'<div class="sheet-hd"><div class="l">盘点'+esc(p.warehouse)+' '+esc(data.header_date)+'</div>'
      +'<div class="c">第 '+(pi+1)+' 页，共 '+pages+' 页</div>'
      +'<div class="r">'+esc(data.header_right||'有效期统计2027年12月31日前')+'</div></div>'
      +'<table class="ptable">'+colgroup()+titleRows()
      +'<tbody>'+slice.map(dataRow).join('')+'</tbody></table></div>';
  }
  const stage=$('#stage'); stage.innerHTML=html; stage.scrollTop=0;
}

function downloadFile(url, name){
  const a=document.createElement('a'); a.href=url; a.download=name||''; document.body.appendChild(a); a.click(); a.remove();
  showToast(name);
}
function showToast(name){
  const t=$('#toast');
  t.innerHTML='<div class="tt"><span class="ck">✓</span>已开始下载</div>'
    +'<div class="td">文件名：<b>'+esc(name)+'</b><br>保存位置：浏览器的「下载」文件夹（如开启了「下载前询问」，则为你所选位置）</div>';
  t.classList.add('show');
  clearTimeout(toastTimer); toastTimer=setTimeout(()=>t.classList.remove('show'),5200);
}

$('#dlzip').onclick=()=>{ if(data) downloadFile('/download/'+data.token, data.zip_name); };
$('#again').onclick=()=>{
  data=null; chosen=null; fileInput.value=''; $('#filetag').innerHTML='';
  go.disabled=true; $('#rbody').style.display='none'; $('#empty').style.display='flex';
};
function esc(s){return String(s).replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));}

// 页眉文字：应用并重新生成（用已选文件重跑，同步预览与下载）
$('#applyhdr').onclick=()=>{ if(chosen) runProcess(); };
$('#hdr').addEventListener('keydown',e=>{ if(e.key==='Enter'&&chosen) runProcess(); });

// 赫眉 HYMN 品牌 logo（SVG 矢量）
const LOGO_SVG = '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 680 200" role="img" aria-label="赫眉 HYMN">'
  +'<rect x="14" y="40" width="312" height="120" rx="2" fill="none" stroke="#221D17" stroke-width="3"/>'
  +'<text x="170" y="135" text-anchor="middle" font-family="\'Noto Sans SC\',\'PingFang SC\',\'Microsoft YaHei\',sans-serif" font-weight="900" font-size="98" letter-spacing="2" fill="#221D17">赫眉</text>'
  +'<text x="362" y="130" font-family="\'Arial Black\',\'Helvetica Neue\',\'DM Sans\',sans-serif" font-weight="900" font-size="80" letter-spacing="6" fill="#221D17">HYMN</text>'
  +'</svg>';
$('#topbrand').innerHTML=LOGO_SVG;
$('#footlogo').innerHTML=LOGO_SVG;

// ── 金色颗粒雾背景（满屏金色颗粒 + 雾密度 + 鼠标搅动）──
(function(){
  const cv=document.getElementById('bg'), cx=cv.getContext('2d');
  let W,H,grains=[],fog=[],t=0;
  const mouse={x:-9999,y:-9999,px:-9999,py:-9999,vx:0,vy:0,active:false};
  const TONES=['233,201,128','214,181,108','190,156,82','231,141,70'];
  function build(){
    W=cv.width=innerWidth; H=cv.height=innerHeight;
    const gn=Math.min(3000, Math.floor(W*H/760));
    grains=[];
    for(let i=0;i<gn;i++){
      const x=Math.random()*W, y=Math.random()*H;
      grains.push({hx:x,hy:y,x:x,y:y,vx:0,vy:0,
        s:Math.random()*1.5+0.5, a:Math.random()*0.5+0.2,
        tone:TONES[(Math.random()*TONES.length)|0],
        tw:Math.random()*6.283, tsp:Math.random()*0.9+0.4});
    }
    const fn=Math.max(8, Math.floor(W*H/170000));
    fog=[];
    for(let i=0;i<fn;i++) fog.push({x:Math.random()*W,y:Math.random()*H,
      r:Math.random()*230+150, ph:Math.random()*6.283, sp:Math.random()*0.3+0.12,
      tone:TONES[(Math.random()*2)|0]});
  }
  addEventListener('mousemove',e=>{mouse.vx=e.clientX-mouse.px;mouse.vy=e.clientY-mouse.py;
    mouse.px=e.clientX;mouse.py=e.clientY;mouse.x=e.clientX;mouse.y=e.clientY;mouse.active=true;});
  addEventListener('mouseout',()=>{mouse.active=false});
  function tick(){
    t+=0.016; mouse.vx*=0.9; mouse.vy*=0.9;
    cx.clearRect(0,0,W,H);
    // 雾密度层
    cx.globalCompositeOperation='lighter';
    for(const f of fog){
      const fx=f.x+Math.cos(t*f.sp+f.ph)*44, fy=f.y+Math.sin(t*f.sp*1.1+f.ph)*44;
      const g=cx.createRadialGradient(fx,fy,0,fx,fy,f.r);
      g.addColorStop(0,'rgba('+f.tone+',0.05)'); g.addColorStop(1,'rgba('+f.tone+',0)');
      cx.fillStyle=g; cx.beginPath(); cx.arc(fx,fy,f.r,0,6.2832); cx.fill();
    }
    // 金色颗粒层
    cx.globalCompositeOperation='source-over';
    for(const p of grains){
      if(mouse.active){
        const dx=p.x-mouse.x, dy=p.y-mouse.y, d2=dx*dx+dy*dy;
        if(d2<15000){ const d=Math.sqrt(d2)||1, ff=1-d/122;
          p.vx+=(dx/d)*ff*2.8 + mouse.vx*ff*0.08;
          p.vy+=(dy/d)*ff*2.8 + mouse.vy*ff*0.08; }
      }
      p.vx+=(p.hx-p.x)*0.02; p.vy+=(p.hy-p.y)*0.02;
      p.vx*=0.8; p.vy*=0.8; p.x+=p.vx; p.y+=p.vy;
      const tw=0.5+0.5*Math.sin(t*p.tsp+p.tw);
      cx.fillStyle='rgba('+p.tone+','+(p.a*tw).toFixed(3)+')';
      cx.fillRect(p.x, p.y, p.s, p.s);
    }
    requestAnimationFrame(tick);
  }
  addEventListener('resize',build); build(); tick();
})();
</script>
</body>
</html>"""


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5001, debug=False)
